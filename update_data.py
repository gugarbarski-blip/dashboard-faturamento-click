"""
update_data.py
--------------
Scrapa o painel admin da Click Impresso, atualiza o Excel com dados do mês atual,
regenera o index.html e faz push para o GitHub.

Uso:
    python update_data.py

Variáveis de ambiente necessárias (crie um arquivo .env ou configure no Windows):
    CLICK_EMAIL   - email de login do painel
    CLICK_SENHA   - senha do painel
    GITHUB_TOKEN  - token do GitHub com acesso ao repositório
"""

import os
import sys
import subprocess
from datetime import date, datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Configurações (altere se necessário)
# ---------------------------------------------------------------------------

ADMIN_LOGIN_URL   = "https://www.clickimpresso.com.br/admin/backup.php"
ADMIN_FORM_ACTION = "https://www.clickimpresso.com.br/admin/logon_confirma.php"
FATURAMENTO_URL   = "https://www.clickimpresso.com.br/admin/faturamento.php"

GITHUB_REPO = "gugarbarski-blip/dashboard-faturamento-click"
BRANCH      = "main"

# Caminho para o repositório local (o script deve estar DENTRO do repositório)
REPO_DIR   = Path(__file__).parent.resolve()
EXCEL_FILE = REPO_DIR / "Faturamento_ClickImpresso.xlsx"
HTML_FILE  = REPO_DIR / "index.html"

# Mapeamento: nome da coluna no site → nome da Loja no Excel
# Ajuste se o site usar nomes diferentes
COLUMN_MAP = {
    "TOTAL":          "TOTAL",
    "Site":           "Site/Protasio",
    "Protasio":       "Site/Protasio",   # combinado no Excel
    "Total":          "Shopping Total",  # "Total" no site = "Shopping Total" no Excel
    "Shopping Total": "Shopping Total",
    "Nilo":           "Nilo",
    "Wenceslau":      "Wenceslau",
    "AssisBrasil":    "AssisBrasil",
    "Andradas":       "Andradas",
    "Anita":          "Anita",
    "Goethe":         "Goethe",
    "Canoas":         "Canoas",
    "Ipiranga":       "Ipiranga",
    "Cachoeirinha":   "Cachoeirinha",
    "Azenha":         "Azenha",
    "Osvaldo":        "Osvaldo",
    "Gravataí":       "Gravataí",
    "Novo Hamburgo":  "Novo Hamburgo",
    "Otavio":         "Otavio",
    "São Leopoldo":   "São Leopoldo",
    "Cristóvão":      "Cristóvão",
}

# ---------------------------------------------------------------------------
# Leitura de credenciais
# ---------------------------------------------------------------------------

def get_env(key):
    val = os.environ.get(key)
    if not val:
        # tenta ler de um arquivo .env simples no mesmo diretório
        env_file = REPO_DIR / ".env"
        if env_file.exists():
            for line in env_file.read_text(encoding="utf-8").splitlines():
                if line.startswith(key + "="):
                    val = line.split("=", 1)[1].strip()
                    break
    if not val:
        raise RuntimeError(
            f"Variável de ambiente '{key}' não encontrada.\n"
            f"Crie o arquivo .env no diretório do script com:\n"
            f"  {key}=seu_valor"
        )
    return val


# ---------------------------------------------------------------------------
# Scraping com Playwright
# ---------------------------------------------------------------------------

def scrape_faturamento():
    """
    Faz login no painel e retorna dict:
      { 'YYYY-MM': { 'NomeLoja': {'qtd': int, 'valor': float} } }
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("Playwright não instalado. Execute: pip install playwright && playwright install chromium")
        sys.exit(1)

    email = get_env("CLICK_EMAIL")
    senha = get_env("CLICK_SENHA")

    today = date.today()
    dt_inicial = "01/01/2000"
    dt_final   = today.strftime("%d/%m/%Y")

    resultado = {}

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page    = context.new_page()

        # 1. Login
        print(f"[1/4] Fazendo login em {ADMIN_LOGIN_URL} ...")
        page.goto(ADMIN_LOGIN_URL)
        page.fill('input[name="mail"]', email)
        page.fill('input[name="senha"]', senha)
        page.click('input[type="submit"]')
        page.wait_for_url("**/admin/**", timeout=15000)
        print(f"      Login OK — URL: {page.url}")

        # 2. Navega para faturamento (tenta direto primeiro, depois via menu)
        print("[2/4] Navegando para faturamento ...")
        page.goto(FATURAMENTO_URL)
        page.wait_for_load_state("domcontentloaded", timeout=15000)
        print(f"      URL atual: {page.url}")
        print(f"      Título: {page.title()}")

        # Identifica o frame correto (pode ser main frame ou um iframe)
        all_frames = page.frames
        print(f"      Frames disponíveis: {len(all_frames)}")
        for i, f in enumerate(all_frames):
            print(f"        [{i}] name='{f.name}' url={f.url}")

        # Tenta encontrar o formulário em qualquer frame
        target_frame = None
        for f in all_frames:
            try:
                f.wait_for_selector('input[name="dt_inicial"]', timeout=3000)
                target_frame = f
                print(f"      Formulário encontrado no frame: name='{f.name}' url={f.url}")
                break
            except Exception:
                pass

        # Se não achou, tenta via menu.php
        if target_frame is None:
            print("      Formulário não encontrado direto. Tentando via menu.php ...")
            page.goto("https://www.clickimpresso.com.br/admin/menu.php?tela=faturamento")
            page.wait_for_load_state("domcontentloaded", timeout=15000)
            import time; time.sleep(2)
            all_frames = page.frames
            print(f"      Frames após menu.php: {len(all_frames)}")
            for i, f in enumerate(all_frames):
                print(f"        [{i}] name='{f.name}' url={f.url}")
            for f in all_frames:
                try:
                    f.wait_for_selector('input[name="dt_inicial"]', timeout=3000)
                    target_frame = f
                    print(f"      Formulário no frame: name='{f.name}' url={f.url}")
                    break
                except Exception:
                    pass

        if target_frame is None:
            # Debug: salva HTML da página para diagnóstico
            html = page.content()
            debug_path = REPO_DIR / "debug_page.html"
            debug_path.write_text(html, encoding="utf-8")
            print(f"      HTML salvo em {debug_path} para diagnóstico")
            print(f"      HTML (500 chars): {html[:500]}")
            raise RuntimeError("Formulário de faturamento não encontrado em nenhum frame")

        # 3. Preenche e submete o formulário correto (o que contém dt_inicial)
        print("[3/4] Buscando dados de faturamento ...")
        target_frame.fill('input[name="dt_inicial"]', dt_inicial)
        target_frame.fill('input[name="dt_final"]',   dt_final)
        try:
            target_frame.select_option('select[name="agrupamento"]', value="mensal")
        except Exception:
            print("      AVISO: select agrupamento não encontrado, continuando...")

        # Submete o form específico que contém dt_inicial (evita clicar no form errado)
        submitted = target_frame.evaluate("""() => {
            const dtInput = document.querySelector('input[name="dt_inicial"]');
            if (!dtInput) return 'dt_inicial nao encontrado';
            const form = dtInput.closest('form');
            if (!form) return 'form pai nao encontrado';
            form.submit();
            return 'ok: ' + (form.id || form.name || 'form sem id');
        }""")
        print(f"      Submit: {submitted}")
        target_frame.wait_for_load_state("networkidle", timeout=30000)

        # Debug: lista cabeçalhos de todas as tabelas para identificar a correta
        tabelas_debug = target_frame.evaluate("""() => {
            const tables = Array.from(document.querySelectorAll('table'));
            return tables.map((t, i) => {
                const rows = t.querySelectorAll('tr');
                if (!rows.length) return {i, headers: [], rowCount: 0};
                const firstRow = rows[0];
                const cells = Array.from(firstRow.querySelectorAll('th, td'));
                const headers = cells.map(c => c.innerText.trim().substring(0, 30));
                return {i, headers: headers.slice(0, 15), rowCount: rows.length};
            });
        }""")
        print(f"      Tabelas encontradas: {len(tabelas_debug)}")
        for t in tabelas_debug:
            print(f"        Tabela[{t['i']}] rows={t['rowCount']} headers={t['headers']}")

        # Debug: mostra primeiras 5 linhas da tabela TOTAL (índice 2)
        sample_rows = target_frame.evaluate("""() => {
            const tables = Array.from(document.querySelectorAll('table'));
            if (tables.length < 3) return [];
            const t = tables[2];
            const rows = Array.from(t.querySelectorAll('tr'));
            return rows.slice(0, 6).map(r => {
                const cells = Array.from(r.querySelectorAll('td, th'));
                return cells.map(c => c.innerText.trim());
            });
        }""")
        print(f"      Primeiras linhas da Tabela[2] (TOTAL): {sample_rows}")

        # 4. Faz o parse da tabela
        print("[4/4] Parseando tabela ...")
        sys.stdout.flush()
        resultado = _parse_via_js(target_frame)

        browser.close()

    return resultado


def _parse_via_js(frame):
    """
    Extrai dados usando API Python do Playwright (sem JS evaluation).
    Usa query_selector_all igual ao código de debug que funcionou.
    Estrutura: cada loja tem tabela própria.
      Linha 0: [NomeLoja]
      Linhas 1+: [qtd(var%), MM/YY, R$valor(var%)]
    """
    import re

    def parse_valor(txt):
        txt = re.sub(r'[R$\s]', '', txt)
        txt = re.sub(r'\(.*?\)', '', txt)
        txt = txt.replace('.', '').replace(',', '.').strip()
        try:
            return round(float(txt), 2)
        except Exception:
            return 0.0

    def parse_qtd(txt):
        m = re.match(r'([\d.]+)', txt.replace(',', ''))
        if m:
            return int(m.group(1).replace('.', ''))
        return 0

    def parse_var(txt):
        m = re.search(r'\(([+-]?\d+)%\)', txt)
        return int(m.group(1)) if m else None

    def parse_periodo(txt):
        # "06/26" → "2026-06"
        m = re.fullmatch(r'(\d{2})/(\d{2})', txt.strip())
        if m:
            return f"20{m.group(2)}-{m.group(1)}"
        return None

    resultado = {}
    lojas_encontradas = []

    tables = frame.query_selector_all('table')
    print(f"       [parse] {len(tables)} tabelas encontradas")
    sys.stdout.flush()

    for table in tables:
        rows = table.query_selector_all('tr')
        if len(rows) < 2:
            continue

        # Primeira linha: deve ter exatamente 1 <td> ou <th> com o nome da loja
        first_row_cells = rows[0].query_selector_all('td, th')
        if len(first_row_cells) != 1:
            continue

        loja_site = first_row_cells[0].inner_text().strip()
        if not loja_site or len(loja_site) > 40:
            continue

        loja_excel = COLUMN_MAP.get(loja_site)
        if not loja_excel:
            continue  # ignora tabelas fora do mapa (ex: ESTOQUE Impresul)

        lojas_encontradas.append(loja_site)

        for row in rows[1:]:
            cells = row.query_selector_all('td')
            if len(cells) < 3:
                continue
            texts = [c.inner_text().strip() for c in cells[:3]]

            periodo = parse_periodo(texts[1])
            if not periodo:
                continue
            if 'R$' not in texts[2]:
                continue

            qtd   = parse_qtd(texts[0])
            var   = parse_var(texts[0]) or parse_var(texts[2])
            valor = parse_valor(texts[2])

            if periodo not in resultado:
                resultado[periodo] = {}

            if loja_excel not in resultado[periodo]:
                resultado[periodo][loja_excel] = {"qtd": qtd, "valor": valor, "var": var}
            else:
                # Site + Protasio são somados em "Site/Protasio"
                resultado[periodo][loja_excel]["qtd"]   += qtd
                resultado[periodo][loja_excel]["valor"]  = round(
                    resultado[periodo][loja_excel]["valor"] + valor, 2)

    print(f"       Lojas mapeadas: {lojas_encontradas}")
    print(f"       Períodos extraídos: {len(resultado)}")
    sys.stdout.flush()
    return resultado


# ---------------------------------------------------------------------------
# Atualização do Excel
# ---------------------------------------------------------------------------

def update_excel(dados_site):
    """
    Atualiza o Excel com os dados do site para o mês atual.
    dados_site: { 'YYYY-MM': { 'NomeLoja': {'qtd': int, 'valor': float, 'var': int|None} } }
    """
    import openpyxl

    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"Excel não encontrado: {EXCEL_FILE}")

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Faturamento"]

    # Colunas: Loja, Tipo, Periodo, Periodo_ISO, Ano, Mes, Quantidade, Valor_RS, Variacao_Pct
    COL = {
        "Loja": 1, "Tipo": 2, "Periodo": 3, "Periodo_ISO": 4,
        "Ano": 5, "Mes": 6, "Quantidade": 7, "Valor_RS": 8, "Variacao_Pct": 9
    }

    today = date.today()
    periodo_atual = today.strftime("%Y-%m")
    mes_atual     = today.month
    ano_atual     = today.year
    periodo_fmt   = today.strftime("%m/%y")  # ex: "06/26"

    if periodo_atual not in dados_site:
        print(f"AVISO: Período {periodo_atual} não encontrado nos dados do site")
        # Usa o período mais recente disponível
        if dados_site:
            periodo_atual = sorted(dados_site.keys())[-1]
            print(f"       Usando período mais recente: {periodo_atual}")
        else:
            print("       Nenhum dado disponível para atualizar")
            return

    dados_mes = dados_site[periodo_atual]

    # Índice de linhas existentes: (Loja, Periodo_ISO) → número da linha
    linhas_existentes = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        loja    = row[COL["Loja"] - 1].value
        periodo = row[COL["Periodo_ISO"] - 1].value
        if loja and periodo:
            linhas_existentes[(loja, periodo)] = row[0].row

    atualizados = 0
    adicionados = 0

    for nome_loja, vals in dados_mes.items():
        qtd   = vals.get('qtd', 0)
        valor = round(vals.get('valor', 0.0), 2)
        var   = vals.get('var')

        # Determina Tipo
        tipo = "Total_Geral" if nome_loja == "TOTAL" else "Loja"

        chave = (nome_loja, periodo_atual)

        if chave in linhas_existentes:
            # Atualiza linha existente
            row_num = linhas_existentes[chave]
            ws.cell(row=row_num, column=COL["Quantidade"]).value  = qtd
            ws.cell(row=row_num, column=COL["Valor_RS"]).value    = valor
            ws.cell(row=row_num, column=COL["Variacao_Pct"]).value = var
            atualizados += 1
        else:
            # Adiciona nova linha
            new_row = [
                nome_loja, tipo, periodo_fmt, periodo_atual,
                ano_atual, mes_atual, qtd, valor, var
            ]
            ws.append(new_row)
            adicionados += 1

    wb.save(EXCEL_FILE)
    print(f"Excel atualizado: {atualizados} linhas atualizadas, {adicionados} adicionadas")
    print(f"Arquivo salvo: {EXCEL_FILE}")


# ---------------------------------------------------------------------------
# Git push
# ---------------------------------------------------------------------------

def git_push(mensagem: str):
    token = get_env("GITHUB_TOKEN")
    remote_url = f"https://{token}@github.com/{GITHUB_REPO}.git"

    cmds = [
        ["git", "-C", str(REPO_DIR), "config", "user.email", "automacao@click.com.br"],
        ["git", "-C", str(REPO_DIR), "config", "user.name",  "Automação Dashboard"],
        ["git", "-C", str(REPO_DIR), "add",    "Faturamento_ClickImpresso.xlsx", "index.html"],
        ["git", "-C", str(REPO_DIR), "commit", "-m", mensagem],
        ["git", "-C", str(REPO_DIR), "push",   remote_url, BRANCH],
    ]

    for cmd in cmds:
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode != 0 and "nothing to commit" not in result.stdout:
            # commit vazio não é erro
            if result.returncode != 0 and "nothing to commit" not in result.stderr:
                print(f"Git: {' '.join(cmd[3:])}")
                print(f"  stdout: {result.stdout.strip()}")
                print(f"  stderr: {result.stderr.strip()}")

    print("Push para GitHub concluído")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    hoje = date.today().strftime("%d/%m/%Y")
    print(f"\n{'='*60}")
    print(f"  Dashboard Click Impresso — Atualização automática")
    print(f"  Data: {hoje}")
    print(f"{'='*60}\n")

    try:
        # 1. Scraping
        dados = scrape_faturamento()

        if not dados:
            print("ERRO: Nenhum dado retornado do scraping. Abortando.")
            sys.exit(1)

        # 2. Atualiza Excel
        update_excel(dados)

        # 3. Regenera HTML
        print("\nRegenerando index.html ...")
        import gerar_html
        gerar_html.gerar(EXCEL_FILE, HTML_FILE)
        print("index.html regenerado com sucesso")

        # 4. Push para GitHub
        print("\nFazendo push para GitHub ...")
        mes = date.today().strftime("%m/%Y")
        git_push(f"auto: atualiza dados {mes}")

        print(f"\nOK: Atualizacao concluida com sucesso em {hoje}")

    except Exception as e:
        print(f"\nERRO: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                           