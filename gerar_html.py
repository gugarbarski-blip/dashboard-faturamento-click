"""
gerar_html.py
-------------
Lê o Excel de faturamento e regenera o const DATA = {...} dentro do index.html.
Preserva todo o HTML/CSS/JS existente — só substitui o bloco de dados.
"""

import json
import re
from pathlib import Path
from datetime import date
from collections import defaultdict


def gerar(excel_path: Path, html_path: Path):
    """
    Ponto de entrada principal.
    Lê o Excel, computa todos os campos de DATA e injeta no index.html.
    """
    dados = _ler_excel(excel_path)
    data_obj = _computar_data(dados)
    _injetar_no_html(html_path, data_obj)


# ---------------------------------------------------------------------------
# 1. Leitura do Excel
# ---------------------------------------------------------------------------

def _ler_excel(path: Path):
    """
    Retorna lista de dicts com as colunas do Excel.
    Filtra apenas linhas com Tipo == 'Loja' (exclui Total_Geral onde necessário).
    """
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb["Faturamento"]

    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        # Sanitiza
        d["Valor_RS"]      = float(d["Valor_RS"] or 0)
        d["Quantidade"]    = int(d["Quantidade"] or 0)
        d["Variacao_Pct"]  = d["Variacao_Pct"]  # pode ser None
        d["Ano"]           = int(d["Ano"] or 0)
        d["Mes"]           = int(d["Mes"] or 0)
        rows.append(d)
    return rows


# ---------------------------------------------------------------------------
# 2. Computação de todos os campos
# ---------------------------------------------------------------------------

def _computar_data(rows):
    lojas_rows  = [r for r in rows if r["Tipo"] == "Loja"]
    total_rows  = [r for r in rows if r["Tipo"] == "Total_Geral"]

    # ── por_loja: soma total por loja (all time) ──────────────────────────
    agg_loja = defaultdict(lambda: {"Valor": 0.0, "Qtd": 0})
    for r in lojas_rows:
        agg_loja[r["Loja"]]["Valor"] += r["Valor_RS"]
        agg_loja[r["Loja"]]["Qtd"]   += r["Quantidade"]
    por_loja = sorted(
        [{"Loja": k, "Valor": round(v["Valor"], 2), "Qtd": v["Qtd"]}
         for k, v in agg_loja.items()],
        key=lambda x: -x["Valor"]
    )

    # ── por_ano: total por ano (Total_Geral) ──────────────────────────────
    agg_ano = defaultdict(float)
    for r in total_rows:
        agg_ano[r["Ano"]] += r["Valor_RS"]
    por_ano = sorted(
        [{"Ano": ano, "Valor_RS": round(v, 2)} for ano, v in agg_ano.items()],
        key=lambda x: x["Ano"]
    )

    # ── recente: últimos 30 meses de totais ───────────────────────────────
    agg_periodo_total = defaultdict(float)
    for r in total_rows:
        agg_periodo_total[r["Periodo_ISO"]] += r["Valor_RS"]
    periodos_sorted = sorted(agg_periodo_total.keys())
    recente = [
        {"Periodo_ISO": p, "Valor_RS": round(agg_periodo_total[p], 2)}
        for p in periodos_sorted[-30:]
    ]

    # ── comp: comparação último ano completo vs penúltimo, por loja ───────
    ano_atual  = date.today().year
    ano_comp1  = ano_atual - 1   # ex: 2025
    ano_comp2  = ano_atual - 2   # ex: 2024
    agg_comp = defaultdict(lambda: {ano_comp1: 0.0, ano_comp2: 0.0})
    for r in lojas_rows:
        if r["Ano"] == ano_comp1:
            agg_comp[r["Loja"]][ano_comp1] += r["Valor_RS"]
        elif r["Ano"] == ano_comp2:
            agg_comp[r["Loja"]][ano_comp2] += r["Valor_RS"]
    comp = sorted(
        [{"Loja": k, str(ano_comp2): round(v[ano_comp2], 2), str(ano_comp1): round(v[ano_comp1], 2)}
         for k, v in agg_comp.items() if v[ano_comp1] > 0],
        key=lambda x: x["Loja"]
    )

    # ── por_ano_loja: por ano e por loja ─────────────────────────────────
    agg_ano_loja = defaultdict(lambda: defaultdict(float))
    for r in lojas_rows:
        agg_ano_loja[str(r["Ano"])][r["Loja"]] += r["Valor_RS"]
    por_ano_loja = {}
    for ano in sorted(agg_ano_loja.keys()):
        lojas_ano = sorted(
            [{"Loja": l, "Valor_RS": round(v, 2)} for l, v in agg_ano_loja[ano].items()],
            key=lambda x: -x["Valor_RS"]
        )
        por_ano_loja[ano] = lojas_ano

    # ── heatmap: valor mensal por loja, últimos 3 anos ────────────────────
    anos_heatmap = [str(ano_atual), str(ano_comp1), str(ano_comp2)]
    heatmap = {}
    for r in lojas_rows:
        ano_s = str(r["Ano"])
        if ano_s not in anos_heatmap:
            continue
        if ano_s not in heatmap:
            heatmap[ano_s] = {}
        loja = r["Loja"]
        if loja not in heatmap[ano_s]:
            heatmap[ano_s][loja] = {}
        mes_s = str(r["Mes"])
        heatmap[ano_s][loja][mes_s] = round(r["Valor_RS"], 2)

    # ── projecao: comparação mensal 2025 vs 2026 ─────────────────────────
    meses_real   = {}  # mes_s → {2025: v, 2026: v}
    for r in total_rows:
        if r["Ano"] in (ano_comp1, ano_atual):
            mes_s = str(r["Mes"])
            if mes_s not in meses_real:
                meses_real[mes_s] = {str(ano_comp1): 0.0, str(ano_atual): 0.0}
            meses_real[mes_s][str(r["Ano"])] += r["Valor_RS"]

    mes_atual = date.today().month
    projecao = {}
    for m in range(1, 13):
        mes_s = str(m)
        v25   = round(meses_real.get(mes_s, {}).get(str(ano_comp1), 0), 2)
        v26   = meses_real.get(mes_s, {}).get(str(ano_atual), 0)
        real_26  = round(v26, 2) if m <= mes_atual else None
        proj_26  = round(v26 * 1.005, 2) if m > mes_atual and v25 > 0 else None
        projecao[mes_s] = {str(ano_comp1): v25, f"{ano_atual}_real": real_26, f"{ano_atual}_proj": proj_26}

    # ── top_var_os: top variação de quantidade 2025→2026 ─────────────────
    # Exclui o mês atual pois pode ser parcial
    qtd_loja_25 = defaultdict(int)
    qtd_loja_26 = defaultdict(int)
    meses_26 = {r["Mes"] for r in lojas_rows if r["Ano"] == ano_atual and r["Mes"] < mes_atual}
    for r in lojas_rows:
        if r["Ano"] == ano_atual and r["Mes"] in meses_26:
            qtd_loja_26[r["Loja"]] += r["Quantidade"]
        elif r["Ano"] == ano_comp1 and r["Mes"] in meses_26:
            qtd_loja_25[r["Loja"]] += r["Quantidade"]
    top_var_os = {}
    for loja in qtd_loja_26:
        q25 = qtd_loja_25.get(loja, 0)
        q26 = qtd_loja_26[loja]
        var = round((q26 / q25 - 1) * 100, 1) if q25 > 0 else None
        if var is not None:
            top_var_os[loja] = {"q2025": q25, "q2026": q26, "var": var}
    top_var_os = dict(sorted(top_var_os.items(), key=lambda x: -x[1]["var"]))

    # ── top_var: top variação de valor 2025→2026 ─────────────────────────
    val_loja_25 = defaultdict(float)
    val_loja_26 = defaultdict(float)
    for r in lojas_rows:
        if r["Ano"] == ano_atual and r["Mes"] in meses_26:
            val_loja_26[r["Loja"]] += r["Valor_RS"]
        elif r["Ano"] == ano_comp1 and r["Mes"] in meses_26:
            val_loja_25[r["Loja"]] += r["Valor_RS"]
    top_var = {}
    for loja in val_loja_26:
        v25 = round(val_loja_25.get(loja, 0), 2)
        v26 = round(val_loja_26[loja], 2)
        var = round((v26 / v25 - 1) * 100, 1) if v25 > 0 else None
        if var is not None:
            top_var[loja] = {"v2025": v25, "v2026": v26, "var": var}
    top_var = dict(sorted(top_var.items(), key=lambda x: -x[1]["var"]))

    # ── qtd_ano: quantidade total por ano ─────────────────────────────────
    agg_qtd_ano = defaultdict(int)
    for r in total_rows:
        agg_qtd_ano[str(r["Ano"])] += r["Quantidade"]
    qtd_ano = dict(sorted(agg_qtd_ano.items(), key=lambda x: -int(x[0])))

    # ── qtd_ano_loja: quantidade por ano e loja ───────────────────────────
    agg_qtd_ano_loja = defaultdict(lambda: defaultdict(int))
    for r in lojas_rows:
        agg_qtd_ano_loja[str(r["Ano"])][r["Loja"]] += r["Quantidade"]
    qtd_ano_loja = {
        ano: dict(sorted(lojas.items(), key=lambda x: -x[1]))
        for ano, lojas in sorted(agg_qtd_ano_loja.items(), key=lambda x: -int(x[0]))
    }

    # ── moy: month over year (valor mensal 2025 vs 2026) ─────────────────
    moy = {}
    for r in lojas_rows:
        if r["Ano"] not in (ano_comp1, ano_atual):
            continue
        loja  = r["Loja"]
        mes_s = str(r["Mes"])
        if loja not in moy:
            moy[loja] = {}
        if mes_s not in moy[loja]:
            moy[loja][mes_s] = {str(ano_comp1): 0.0, str(ano_atual): 0.0}
        moy[loja][mes_s][str(r["Ano"])] += r["Valor_RS"]

    # Calcula variação
    for loja in moy:
        for mes_s in moy[loja]:
            v25 = moy[loja][mes_s].get(str(ano_comp1), 0)
            v26 = moy[loja][mes_s].get(str(ano_atual), 0)
            var = round((v26 / v25 - 1) * 100, 1) if v25 > 0 else None
            moy[loja][mes_s] = {
                str(ano_comp1): round(v25, 2),
                str(ano_atual): round(v26, 2),
                "var": var
            }

    # ── metas: mantém as metas existentes (não altera) ───────────────────
    # As metas são definidas manualmente no HTML original — não as recalculamos
    metas_existentes = _extrair_metas_do_html(
        Path(__file__).parent / "index.html"
    )

    # ── quedas: lojas com 4+ meses negativos no ano atual ────────────────
    quedas = {}
    for loja, meses_loja in moy.items():
        negativos = [(int(m), meses_loja[m]["var"])
                     for m in meses_loja
                     if meses_loja[m].get("var") is not None and meses_loja[m]["var"] < 0]
        if len(negativos) >= 4:
            media = round(sum(v for _, v in negativos) / len(negativos), 1)
            quedas[loja] = {
                "meses_negativos": len(negativos),
                "media_var": media,
                "detalhes": [{"mes": m, "var": round(v, 1)} for m, v in sorted(negativos)]
            }
    quedas = dict(sorted(quedas.items(), key=lambda x: -x[1]["meses_negativos"]))

    # ── ticket_evolucao: ticket médio por loja por ano ────────────────────
    ticket_por_ano_loja = defaultdict(lambda: defaultdict(lambda: {"valor": 0.0, "qtd": 0}))
    for r in lojas_rows:
        if r["Quantidade"] > 0:
            ticket_por_ano_loja[r["Loja"]][str(r["Ano"])]["valor"] += r["Valor_RS"]
            ticket_por_ano_loja[r["Loja"]][str(r["Ano"])]["qtd"]   += r["Quantidade"]

    ticket_evolucao = {}
    principais = ["Nilo","Site/Protasio","Goethe","Wenceslau","Andradas",
                  "Anita","AssisBrasil","Canoas","Ipiranga","Cachoeirinha"]
    for loja in principais:
        if loja in ticket_por_ano_loja:
            ticket_evolucao[loja] = {
                ano: round(v["valor"] / v["qtd"], 2) if v["qtd"] > 0 else None
                for ano, v in sorted(ticket_por_ano_loja[loja].items())
            }

    # ── ticket_anual ──────────────────────────────────────────────────────
    ticket_anual = {}
    for loja, anos in ticket_por_ano_loja.items():
        for ano, v in anos.items():
            if ano not in ticket_anual:
                ticket_anual[ano] = {}
            if v["qtd"] > 0:
                ticket_anual[ano][loja] = round(v["valor"] / v["qtd"], 2)
    ticket_anual = dict(sorted(ticket_anual.items(), key=lambda x: -int(x[0])))

    # ── ticket_mensal ─────────────────────────────────────────────────────
    ticket_mensal_agg = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {"valor": 0.0, "qtd": 0})))
    for r in lojas_rows:
        if r["Quantidade"] > 0 and r["Ano"] >= (ano_atual - 3):
            ticket_mensal_agg[str(r["Ano"])][str(r["Mes"])][r["Loja"]]["valor"] += r["Valor_RS"]
            ticket_mensal_agg[str(r["Ano"])][str(r["Mes"])][r["Loja"]]["qtd"]   += r["Quantidade"]

    ticket_mensal = {}
    for ano, meses_d in sorted(ticket_mensal_agg.items(), key=lambda x: -int(x[0])):
        ticket_mensal[ano] = {}
        for mes, lojas_d in sorted(meses_d.items(), key=lambda x: -int(x[0])):
            ticket_mensal[ano][mes] = {
                loja: round(v["valor"] / v["qtd"], 2)
                for loja, v in lojas_d.items() if v["qtd"] > 0
            }

    # ── ult_mes: variação do último mês disponível ────────────────────────
    ultimo_periodo = sorted({r["Periodo_ISO"] for r in lojas_rows})[-1]
    ano_ult  = int(ultimo_periodo.split("-")[0])
    mes_ult  = int(ultimo_periodo.split("-")[1])
    periodo_anterior = f"{ano_ult - 1}-{mes_ult:02d}"

    val_ult   = {r["Loja"]: r["Valor_RS"] for r in lojas_rows if r["Periodo_ISO"] == ultimo_periodo}
    val_ant   = {r["Loja"]: r["Valor_RS"] for r in lojas_rows if r["Periodo_ISO"] == periodo_anterior}

    ult_mes = []
    for loja in sorted(val_ult.keys()):
        ant = val_ant.get(loja, 0)
        ult = val_ult[loja]
        var = round((ult / ant - 1) * 100) if ant > 0 else None
        ult_mes.append({"Loja": loja, "Variacao_Pct": var})

    # ── Monta objeto final ────────────────────────────────────────────────
    return {
        "por_loja":        por_loja,
        "por_ano":         por_ano,
        "recente":         recente,
        "comp":            comp,
        "por_ano_loja":    por_ano_loja,
        "heatmap":         heatmap,
        "projecao":        projecao,
        "top_var_os":      top_var_os,
        "top_var":         top_var,
        "qtd_ano":         qtd_ano,
        "qtd_ano_loja":    qtd_ano_loja,
        "moy":             moy,
        "metas":           metas_existentes,
        "quedas":          quedas,
        "ticket_evolucao": ticket_evolucao,
        "ticket_anual":    ticket_anual,
        "ticket_mensal":   ticket_mensal,
        "ult_mes":         ult_mes,
    }


def _extrair_metas_do_html(html_path: Path) -> dict:
    """Extrai o campo 'metas' do HTML existente para preservar os valores."""
    if not html_path.exists():
        return {}
    html = html_path.read_text(encoding="utf-8")
    m = re.search(r'metas:\s*(\{[^}]+\})', html)
    if m:
        try:
            return json.loads(m.group(1))
        except:
            pass
    return {}


# ---------------------------------------------------------------------------
# 3. Injeção no HTML
# ---------------------------------------------------------------------------

def _injetar_no_html(html_path: Path, data_obj: dict):
    """
    Substitui o bloco `const DATA = {...};` no index.html pelo novo JSON.
    Preserva todo o restante do arquivo intacto.
    """
    if not html_path.exists():
        raise FileNotFoundError(f"HTML não encontrado: {html_path}")

    html = html_path.read_text(encoding="utf-8")

    # Gera o JSON compacto (sem espaços extras)
    data_json = json.dumps(data_obj, ensure_ascii=False, separators=(",", ":"))

    # Monta o bloco novo com os campos separados por linha (mais legível)
    campos = []
    for chave, valor in data_obj.items():
        campos.append(f"  {chave}: {json.dumps(valor, ensure_ascii=False, separators=(',', ':'))}")
    novo_bloco = "const DATA = {\n" + ",\n".join(campos) + "\n};"

    # Substitui o bloco existente usando regex
    # Captura tudo entre "const DATA = {" e o "}" de fechamento do objeto + ";"
    padrao = re.compile(
        r'const DATA\s*=\s*\{.*?\};',
        re.DOTALL
    )

    if not padrao.search(html):
        raise RuntimeError("Bloco 'const DATA = {...};' não encontrado no HTML")

    novo_html = padrao.sub(novo_bloco, html)

    # Atualiza a linha de "Dados atualizados até"
    hoje = date.today()
    meses_pt = {
        1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",5:"Maio",6:"Junho",
        7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro"
    }
    mes_nome = meses_pt[hoje.month]
    novo_html = re.sub(
        r'Dados atualizados até [^·]+·',
        f"Dados atualizados até {mes_nome}/{hoje.year} ·",
        novo_html
    )

    html_path.write_text(novo_html, encoding="utf-8")
    print(f"index.html atualizado: {html_path}")


if __name__ == "__main__":
    import sys
    repo = Path(__file__).parent
    gerar(repo / "Faturamento_ClickImpresso.xlsx", repo / "index.html")
    print("Concluído.")
